#!/usr/bin/env python3
"""Build the auditable BPSC biology PYQ dataset used by the rearranged notes.

Inputs:
  EDUTERIA BIO NOTES/56-59th BPSC Preliminary Examination Question Paper.xlsx
  EDUTERIA BIO NOTES/BPSC_Master_Question_Bank_60_to_71.xlsx

Output:
  build_biology/pyq_analysis.csv
  build_biology/pyq_summary.json

The scope intentionally has two labels:
* core: biology concepts and direct life-science applications
* extended: ecology, conservation, health programmes and agri-biology

Requires openpyxl (see scripts/requirements_biology_notes.txt).
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BIO = ROOT / "EDUTERIA BIO NOTES"
OUT = ROOT / "build_biology"

# q_no: (topic, scope). Selection was reviewed question-by-question; it is not
# a keyword-only classification. This avoids false matches such as "cell" in
# "photoelectric cell" and "body" in polity questions.
SELECTION: dict[str, dict[int, tuple[str, str]]] = {
    "56-59th BPSC (2015)": {
        19: ("Ecology, Environment & Conservation", "extended"),
        20: ("Ecology, Environment & Conservation", "extended"),
        41: ("Microbes, Disease & Immunity", "core"),
        42: ("Nutrition & Biomolecules", "core"),
        43: ("Microbes, Disease & Immunity", "core"),
        44: ("Genetics, DNA & Biotechnology", "core"),
        45: ("Animal Diversity & Applied Biology", "core"),
        47: ("Plant Form, Function & Classification", "core"),
        48: ("Excretion & Osmoregulation", "core"),
        49: ("Circulation, Blood & Immunity", "core"),
        50: ("Cell Biology", "core"),
        134: ("Plant Form, Function & Classification", "core"),
        135: ("Microbes, Disease & Immunity", "core"),
        137: ("Genetics, DNA & Biotechnology", "core"),
        142: ("Microbes, Disease & Immunity", "core"),
    },
    "60-62nd BPSC (Combined)": {
        31: ("Nutrition & Biomolecules", "core"),
        33: ("Genetics, DNA & Biotechnology", "core"),
        34: ("Nutrition & Biomolecules", "core"),
        35: ("Nutrition & Biomolecules", "core"),
        37: ("Microbes, Disease & Immunity", "core"),
        39: ("Endocrine & Reproduction", "core"),
        43: ("Endocrine & Reproduction", "core"),
        45: ("Nutrition & Biomolecules", "core"),
        46: ("Nutrition & Biomolecules", "core"),
        48: ("Ecology, Environment & Conservation", "extended"),
        50: ("Microbes, Disease & Immunity", "core"),
    },
    "63rd BPSC": {
        87: ("Excretion & Osmoregulation", "core"),
        91: ("Endocrine & Reproduction", "core"),
        92: ("Microbes, Disease & Immunity", "core"),
        93: ("Nutrition & Biomolecules", "core"),
    },
    "64th BPSC": {
        5: ("Genetics, DNA & Biotechnology", "core"),
        6: ("Nutrition & Biomolecules", "core"),
        7: ("Nervous System & Senses", "core"),
        8: ("Nervous System & Senses", "core"),
        13: ("Nutrition & Biomolecules", "core"),
        15: ("Ecology, Environment & Conservation", "extended"),
        17: ("Cell Biology", "core"),
        18: ("Microbes, Disease & Immunity", "core"),
    },
    "65th BPSC": {
        96: ("Ecology, Environment & Conservation", "extended"),
        99: ("Nutrition & Biomolecules", "core"),
        101: ("Genetics, DNA & Biotechnology", "core"),
        102: ("Ecology, Environment & Conservation", "core"),
        103: ("Ecology, Environment & Conservation", "core"),
        104: ("Plant Form, Function & Classification", "core"),
        105: ("Plant Form, Function & Classification", "core"),
        106: ("Skeleton, Joints & Muscle", "core"),
        107: ("Microbes, Disease & Immunity", "core"),
        108: ("Circulation, Blood & Immunity", "core"),
        109: ("Endocrine & Reproduction", "core"),
        110: ("Microbes, Disease & Immunity", "core"),
        118: ("Nervous System & Senses", "core"),
    },
    "66th BPSC": {
        1: ("Skeleton, Joints & Muscle", "core"),
        11: ("Nutrition & Biomolecules", "core"),
        13: ("Nutrition & Biomolecules", "core"),
        21: ("Nutrition & Biomolecules", "core"),
        22: ("Endocrine & Reproduction", "core"),
        23: ("Nervous System & Senses", "core"),
        24: ("Endocrine & Reproduction", "core"),
        25: ("Digestion", "core"),
        26: ("Plant Form, Function & Classification", "core"),
        27: ("Nutrition & Biomolecules", "core"),
        28: ("Plant Form, Function & Classification", "core"),
        29: ("Nutrition & Biomolecules", "core"),
        30: ("Plant Form, Function & Classification", "core"),
    },
    "67th BPSC": {
        55: ("Plant Form, Function & Classification", "core"),
        56: ("Plant Form, Function & Classification", "core"),
        57: ("Plant Form, Function & Classification", "core"),
        58: ("Plant Form, Function & Classification", "core"),
        59: ("Plant Form, Function & Classification", "core"),
        60: ("Microbes, Disease & Immunity", "core"),
        61: ("Circulation, Blood & Immunity", "core"),
        62: ("Digestion", "core"),
        63: ("Nutrition & Biomolecules", "core"),
        64: ("Microbes, Disease & Immunity", "core"),
        93: ("Ecology, Environment & Conservation", "extended"),
        94: ("Ecology, Environment & Conservation", "extended"),
        100: ("Ecology, Environment & Conservation", "extended"),
        106: ("Microbes, Disease & Immunity", "extended"),
        148: ("Ecology, Environment & Conservation", "extended"),
    },
    "68th BPSC": {
        66: ("Animal Diversity & Applied Biology", "extended"),
        70: ("Nutrition & Biomolecules", "extended"),
        97: ("Ecology, Environment & Conservation", "extended"),
        123: ("Ecology, Environment & Conservation", "extended"),
        131: ("Plant Form, Function & Classification", "core"),
        132: ("Plant Form, Function & Classification", "core"),
        133: ("Plant Form, Function & Classification", "core"),
        134: ("Plant Form, Function & Classification", "core"),
        135: ("Cell Biology", "core"),
        136: ("Skeleton, Joints & Muscle", "core"),
        137: ("Ecology, Environment & Conservation", "core"),
        138: ("Animal Diversity & Applied Biology", "core"),
        139: ("Circulation, Blood & Immunity", "core"),
        140: ("Skeleton, Joints & Muscle", "core"),
    },
    "69th BPSC": {
        3: ("Animal Diversity & Applied Biology", "core"),
        4: ("Microbes, Disease & Immunity", "core"),
        7: ("Nervous System & Senses", "core"),
        8: ("Microbes, Disease & Immunity", "core"),
        18: ("Microbes, Disease & Immunity", "core"),
        20: ("Genetics, DNA & Biotechnology", "core"),
        22: ("Microbes, Disease & Immunity", "core"),
        25: ("Genetics, DNA & Biotechnology", "core"),
        30: ("Nervous System & Senses", "core"),
        46: ("Ecology, Environment & Conservation", "extended"),
        47: ("Ecology, Environment & Conservation", "extended"),
        53: ("Plant Form, Function & Classification", "extended"),
        54: ("Ecology, Environment & Conservation", "extended"),
        86: ("Ecology, Environment & Conservation", "extended"),
        87: ("Ecology, Environment & Conservation", "extended"),
        93: ("Ecology, Environment & Conservation", "extended"),
    },
    "70th BPSC": {
        3: ("Digestion", "core"),
        31: ("Plant Form, Function & Classification", "core"),
        39: ("Ecology, Environment & Conservation", "extended"),
        41: ("Microbes, Disease & Immunity", "core"),
        45: ("Skeleton, Joints & Muscle", "core"),
        56: ("Plant Form, Function & Classification", "core"),
        66: ("Circulation, Blood & Immunity", "core"),
        79: ("Cell Biology", "core"),
        120: ("Plant Form, Function & Classification", "core"),
        134: ("Digestion", "core"),
        143: ("Ecology, Environment & Conservation", "core"),
        148: ("Ecology, Environment & Conservation", "extended"),
    },
    "71st BPSC (2025)": {
        14: ("Ecology, Environment & Conservation", "extended"),
        40: ("Microbes, Disease & Immunity", "extended"),
        112: ("Ecology, Environment & Conservation", "extended"),
        114: ("Ecology, Environment & Conservation", "extended"),
        115: ("Animal Diversity & Applied Biology", "extended"),
        134: ("Microbes, Disease & Immunity", "core"),
        135: ("Endocrine & Reproduction", "core"),
        138: ("Genetics, DNA & Biotechnology", "core"),
        139: ("Endocrine & Reproduction", "core"),
        140: ("Microbes, Disease & Immunity", "core"),
        141: ("Circulation, Blood & Immunity", "core"),
        142: ("Cell Biology", "core"),
        143: ("Genetics, DNA & Biotechnology", "core"),
        144: ("Ecology, Environment & Conservation", "core"),
        145: ("Plant Form, Function & Classification", "extended"),
        146: ("Plant Form, Function & Classification", "core"),
        147: ("Plant Form, Function & Classification", "core"),
        148: ("Plant Form, Function & Classification", "core"),
        149: ("Ecology, Environment & Conservation", "extended"),
        150: ("Ecology, Environment & Conservation", "core"),
    },
}

QUESTION_OVERRIDES = {
    ("56-59th BPSC (2015)", 42): "Energy is stored in liver and muscles in the form of:",
}

# Content answers for unanswered 71st rows and explicit correction/ambiguity notes.
STUDY_OVERRIDES: dict[tuple[str, int], tuple[str, str, str]] = {
    ("56-59th BPSC (2015)", 42): ("Glycogen", "restored", "Question text restored from the repository's curated biology bank; the XLSX row has missing text."),
    ("56-59th BPSC (2015)", 44): ("Adaptation (exam-key answer)", "key-sensitive", "Adaptation is the expected BPSC answer; mutation/recombination generate heritable variation, on which selection acts."),
    ("56-59th BPSC (2015)", 134): ("Dried flower buds", "restored", "The workbook option/answer row is truncated; clove is the dried flower bud."),
    ("60-62nd BPSC (Combined)", 35): ("Tocopherol–Beriberi is the mismatched pair", "corrected", "Tocopherol is vitamin E; beriberi is due to thiamine (vitamin B1) deficiency."),
    ("60-62nd BPSC (Combined)", 37): ("Chloramphenicol", "corrected", "The workbook key points to Quinine; chloramphenicol is the antibiotic among the listed choices."),
    ("60-62nd BPSC (Combined)", 39): ("Protein (peptide hormone)", "corrected", "The workbook letter maps to 'None'; insulin is a peptide/protein hormone."),
    ("60-62nd BPSC (Combined)", 48): ("Sulfuric acid is generally the dominant acid; nitric acid also contributes", "ambiguous", "The workbook chooses the combined/none option. The singular word 'maximum' makes H2SO4 the standard response, although both H2SO4 and HNO3 occur."),
    ("60-62nd BPSC (Combined)", 50): ("Cobalt-60 (exam key)", "key-sensitive", "The supplied key selects Co-60; P-32 has also had haematological uses, so follow the official/key context."),
    ("63rd BPSC", 92): ("More than one: penicillin and sulfadiazine are antibacterial drugs", "ambiguous", "The broad term 'antibiotic' makes more than one option defensible; the supplied key selects option E."),
    ("66th BPSC", 25): ("Posterior tongue in the old sensitivity map; all basic tastes are broadly distributed", "key-sensitive", "The classic tongue map is an oversimplification. Posterior sensitivity to bitterness does not mean bitterness is sensed only there."),
    ("67th BPSC", 56): ("Chlorophyll (historical option)", "key-sensitive", "Fungi lack chlorophyll, but modern classification does not regard fungi as plants; they form a separate kingdom."),
    ("69th BPSC", 18): ("A harmless antigenic piece/instruction from SARS-CoV-2", "verified", "Vaccines train immunity without introducing a disease-causing dose of the real virus."),
    ("69th BPSC", 22): ("A modified harmless carrier virus delivers genetic instructions", "verified", "This is the defining idea of a viral-vector vaccine."),
    ("70th BPSC", 31): ("Nitrogen (wording-dependent)", "key-sensitive", "The workbook selects nitrogen for early vegetative growth; phosphorus is especially associated with early root development."),
    ("70th BPSC", 56): ("More than one listed organism can form spores", "ambiguous", "Fungi, many algae, yeast and ferns can all form spores in relevant life cycles; the workbook's single choice is defective."),
    ("71st BPSC (2025)", 14): ("BirdLife International", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 40): ("None of the statements is incorrect", "added-answer", "All three listed U-WIN features are supported; answer absent in supplied workbook."),
    ("71st BPSC (2025)", 112): ("Madhya Pradesh", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 114): ("Valmiki National Park / Valmiki Tiger Reserve", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 115): ("Bull/Ox (exam-oriented answer)", "key-sensitive", "Secondary lists conflict between Bull/Ox and Gaur; retain the exam-oriented option and revise from the official key used for your paper set."),
    ("71st BPSC (2025)", 134): ("Platinum compound (e.g., cisplatin)", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 135): ("Levonorgestrel", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 138): ("Poly(3-hydroxybutyrate-co-3-hydroxyvalerate)", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 139): ("Endocrine glands", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 140): ("Parasitism; more specifically intracellular endoparasitism", "ambiguous", "Options A and D overlap; answer absent in supplied workbook."),
    ("71st BPSC (2025)", 141): ("IgM", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 142): ("Lipids and proteins in species/cell-specific proportions", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 143): ("UAA", "added-answer", "UAA is one of the three stop/nonsense codons; answer absent in workbook."),
    ("71st BPSC (2025)", 144): ("Endangered", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 145): ("Sugarcane", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 146): ("Transport of water and dissolved minerals", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 147): ("A. P. de Candolle", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 148): ("Mesocarp", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 149): ("16 September", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
    ("71st BPSC (2025)", 150): ("Increased mineral absorption and protection from disease", "added-answer", "Answer absent in supplied workbook; content answer added by editor."),
}


def load_rows() -> dict[tuple[str, int], dict]:
    rows: dict[tuple[str, int], dict] = {}

    wb = load_workbook(BIO / "56-59th BPSC Preliminary Examination Question Paper.xlsx", data_only=True, read_only=True)
    ws = wb.active
    ws.calculate_dimension(force=True)  # this workbook intentionally lacks a dimension record
    for r in ws.iter_rows(min_row=5, max_row=180, values_only=True):
        if isinstance(r[0], (int, float)) and r[1]:
            qno = int(r[0])
            rows[("56-59th BPSC (2015)", qno)] = {
                "question": str(r[1]),
                "options": [str(x) if x is not None else "" for x in r[2:6]],
                "source_key": str(r[6] or ""),
            }

    wb = load_workbook(BIO / "BPSC_Master_Question_Bank_60_to_71.xlsx", data_only=True, read_only=True)
    ws = wb["Master_All_Questions"]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[0] and isinstance(r[1], (int, float)) and r[2]:
            exam, qno = str(r[0]), int(r[1])
            rows[(exam, qno)] = {
                "question": str(r[2]),
                "options": [str(x) if x is not None else "" for x in r[3:8]],
                "source_key": str(r[9] or r[8] or ""),
            }
    return rows


def resolve_answer(raw: str, options: list[str]) -> str:
    s = raw.strip()
    m = re.fullmatch(r"\(?([A-Ea-e])\)?", s)
    if m:
        i = ord(m.group(1).upper()) - ord("A")
        return options[i] if 0 <= i < len(options) and options[i] else s.upper()
    return s


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    source = load_rows()
    records = []
    missing = []
    for exam, qmap in SELECTION.items():
        for qno, (topic, scope) in qmap.items():
            row = source.get((exam, qno))
            if not row:
                missing.append((exam, qno))
                continue
            question = QUESTION_OVERRIDES.get((exam, qno), row["question"])
            source_answer = resolve_answer(row["source_key"], row["options"])
            if (exam, qno) in STUDY_OVERRIDES:
                study_answer, status, note = STUDY_OVERRIDES[(exam, qno)]
            else:
                study_answer, status, note = source_answer, "source-key", "Answer follows the verified/source answer field in the supplied workbook."
            records.append({
                "exam": exam,
                "q_no": qno,
                "scope": scope,
                "topic": topic,
                "question": " ".join(question.split()),
                "options": " | ".join(x for x in row["options"] if x),
                "source_key_raw": row["source_key"],
                "source_answer": source_answer,
                "study_answer": study_answer,
                "status": status,
                "editor_note": note,
            })
    if missing:
        raise SystemExit(f"Missing selected rows: {missing}")

    fields = list(records[0])
    with (OUT / "pyq_analysis.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        w.writeheader(); w.writerows(records)

    topic = Counter(r["topic"] for r in records)
    core_topic = Counter(r["topic"] for r in records if r["scope"] == "core")
    exam = Counter(r["exam"] for r in records)
    scope = Counter(r["scope"] for r in records)
    years_by_topic = defaultdict(set)
    for r in records:
        years_by_topic[r["topic"]].add(r["exam"])
    ranking = []
    for rank, (name, n) in enumerate(sorted(topic.items(), key=lambda x: (-x[1], x[0])), 1):
        ranking.append({
            "rank": rank,
            "topic": name,
            "all_questions": n,
            "core_questions": core_topic[name],
            "papers_present": len(years_by_topic[name]),
            "priority": "A" if n >= 14 or len(years_by_topic[name]) >= 8 else ("B" if n >= 7 else "C"),
        })
    summary = {
        "method": "Manual question-by-question classification of the two supplied XLSX files; direct biology is core, ecology/conservation/public-health/agri-biology is extended.",
        "total_questions": len(records),
        "scope_counts": dict(scope),
        "exam_counts": dict(exam),
        "topic_ranking": ranking,
        "status_counts": dict(Counter(r["status"] for r in records)),
    }
    (OUT / "pyq_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(records)} reviewed questions")
    for r in ranking:
        print(f"{r['rank']:2}. {r['topic']:<45} {r['all_questions']:2} questions | {r['papers_present']:2} papers | Priority {r['priority']}")


if __name__ == "__main__":
    main()
