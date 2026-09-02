#!/usr/bin/env python3
"""Build the auditable BPSC *physics* PYQ dataset used by the consolidated
physics notes (BPSC_Physics_Consolidated_Notes_A4.pdf).

Inputs (both already in the repository):
  physic notes eduteria/56-59th BPSC Preliminary Examination Question Paper.xlsx
  EDUTERIA BIO NOTES/BPSC_Master_Question_Bank_60_to_71.xlsx   (60-62nd … 71st)

Outputs:
  build_physics/pyq_analysis.csv   – every selected question with source key,
                                     study answer, topic, scope and notes
  build_physics/pyq_summary.json   – topic ranking (all papers + last 6 papers)

Every question of all 11 papers (1,623 rows) was read; the SELECTION below is
a manual, question-by-question classification – not a keyword match.  Two
scopes are used:
  core     – a physics concept that is taught in the Eduteria lecture notes
  extended – physics-adjacent (space/astronomy, atomic structure, science
             history, physics-flavoured current affairs) – useful context but
             not a lecture topic on its own
Pure current-affairs items (launch-vehicle names, robot names, awards) and
quantitative-aptitude "train crossing a bridge" sums are deliberately left out.
The 71st (2025) paper carries no answer key in the workbook, so the study
answers for it were worked out by the editor and are flagged "added-answer".
"""
from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PHY = ROOT / "physic notes eduteria"
BIO = ROOT / "EDUTERIA BIO NOTES"
OUT = ROOT / "build_physics"

EXAM_ORDER = [
    "56-59th BPSC (2015)", "60-62nd BPSC (Combined)", "63rd BPSC", "64th BPSC",
    "65th BPSC", "66th BPSC", "67th BPSC", "68th BPSC", "69th BPSC", "70th BPSC",
    "71st BPSC (2025)",
]
RECENT = EXAM_ORDER[-6:]          # 66th … 71st = the "last 5–6 papers" lens

# Topic labels = the unit titles of the consolidated notes (so the ranking maps
# 1:1 onto the book), plus one extended-only label for space/astronomy.
T_UNITS = "U1 Units, Dimensions & Measuring Instruments"
T_LIGHT1 = "U2 Light I: Nature of Light, Reflection & Mirrors"
T_LIGHT2 = "U3 Light II: Refraction, Lenses & the Human Eye"
T_LIGHT3 = "U4 Light III: Dispersion, Scattering, Wave Optics & Instruments"
T_WAVES = "U5 Waves & the Electromagnetic Spectrum"
T_SOUND = "U6 Sound"
T_HEAT = "U7 Heat, Temperature & Thermometry"
T_THERMO = "U8 Thermal Expansion, Radiation Laws & Thermodynamics"
T_KIN = "U9 Kinematics"
T_NLM = "U10 Laws of Motion, Friction, Work-Power-Energy"
T_GRAV = "U11 Circular Motion & Gravitation"
T_FLUID = "U12 Mechanical Properties of Fluids"
T_ELEC = "U13 Electric Charge, Field & Current Electricity"
T_POWER = "U14 Electrical Power, Devices & Domestic Electricity"
T_MAG = "U15 Magnetism & Electromagnetism"
T_MODERN = "U16 Modern & Nuclear Physics"
T_SPACE = "Space & Astronomy (extended only)"

# exam -> {q_no: (topic, scope)}
SELECTION: dict[str, dict[int, tuple[str, str]]] = {
    "56-59th BPSC (2015)": {
        138: (T_MODERN, "core"),      # Nagasaki bomb – plutonium
        139: (T_GRAV, "core"),        # Copernicus – heliocentric (L13 history list)
        140: (T_MODERN, "core"),      # cosmic rays
        142: (T_MODERN, "extended"),  # ionising radiation causes & cures cancer
    },
    "60-62nd BPSC (Combined)": {
        49: (T_MODERN, "core"),       # God particle – Higgs boson
        50: (T_MODERN, "extended"),   # radio-isotope for leukaemia
        71: (T_MODERN, "extended"),   # LIGO-India (gravitational waves)
    },
    "63rd BPSC": {
        85: (T_LIGHT1, "core"),       # visible spectrum 3900–7600 Å
        89: (T_MODERN, "extended"),   # nucleus – positively charged centre
        90: (T_HEAT, "core"),         # sublimation
        99: (T_POWER, "extended"),    # acid in a car battery (lead-acid, L17)
    },
    "64th BPSC": {
        2: (T_MODERN, "extended"),    # hydrogen has no neutron
        3: (T_UNITS, "core"),         # angstrom – wavelength
        4: (T_UNITS, "core"),         # frequency – hertz
        9: (T_LIGHT3, "core"),        # red λ > violet λ
        10: (T_HEAT, "core"),         # 40 °C = 104 °F
        11: (T_POWER, "core"),        # unit of electric power – watt
        12: (T_POWER, "core"),        # electric motor: electrical → mechanical
        14: (T_POWER, "core"),        # ammeter measures current
        16: (T_MODERN, "core"),       # Einstein Nobel – photoelectric effect
        19: (T_UNITS, "core"),        # hygrometer – humidity
        20: (T_UNITS, "core"),        # unit of pressure N/m²
    },
    "65th BPSC": {
        92: (T_MODERN, "extended"),   # mass number = p + n
        111: (T_UNITS, "core"),       # unit of pressure kg/cm²
        112: (T_LIGHT1, "core"),      # sunlight reaches earth in ≈ 8 min 20 s
        113: (T_KIN, "core"),         # scalar quantity – pressure
        114: (T_UNITS, "core"),       # strain has no unit
        115: (T_SOUND, "core"),       # sound in air – longitudinal
        116: (T_HEAT, "core"),        # 50 °C = 122 °F
        117: (T_UNITS, "core"),       # hertz – frequency
        119: (T_POWER, "core"),       # ammeter
        120: (T_POWER, "core"),       # watt
    },
    "66th BPSC": {
        1: (T_WAVES, "core"),         # infrared for muscle ache
        2: (T_ELEC, "core"),          # parallel resistors 1.403 kΩ
        3: (T_ELEC, "core"),          # semiconductor resistance falls on heating
        4: (T_POWER, "core"),         # Faraday constant – universal constant
        5: (T_UNITS, "core"),         # light-year – distance
        6: (T_LIGHT2, "core"),        # frequency unchanged on refraction
        7: (T_WAVES, "core"),         # EM-wave speed 3 × 10⁸ m/s
        8: (T_SPACE, "extended"),     # first man on the Moon
        9: (T_MODERN, "core"),        # neutrons in ₉₄Pu²⁴² = 148
        10: (T_FLUID, "core"),        # honey – highest viscosity
        20: (T_HEAT, "core"),         # poorest conductor of heat – lead
    },
    "67th BPSC": {
        65: (T_GRAV, "core"),         # washing machine – centrifugation
        66: (T_LIGHT2, "core"),       # light slowest in glass
        67: (T_KIN, "core"),          # speed is not a vector
        68: (T_GRAV, "core"),         # faster spin → weight at equator decreases
        69: (T_KIN, "extended"),      # Galileo first defined speed
        71: (T_MODERN, "extended"),   # theory of relativity – Einstein
        72: (T_THERMO, "core"),       # Thomson (thermoelectric) effect
        73: (T_LIGHT2, "core"),       # magnifying glass – convex lens
        74: (T_MAG, "core"),          # paramagnetic – platinum
        75: (T_MODERN, "extended"),   # nucleus = protons + neutrons
        141: (T_GRAV, "extended"),    # chief heavenly body – Sun
    },
    "68th BPSC": {
        75: (T_MODERN, "extended"),   # LLNL fusion-ignition breakthrough (Dec 2022)
        111: (T_LIGHT2, "core"),      # black strips on convex lens → dimmer image
        112: (T_SOUND, "core"),       # shrillness ← frequency (pitch)
        113: (T_MODERN, "core"),      # photodiode for digital applications
        114: (T_NLM, "core"),         # ball bearings → rolling friction
        115: (T_NLM, "core"),         # goalkeeper – rate of change of momentum
        116: (T_MODERN, "core"),      # alpha ray – positively charged
        117: (T_GRAV, "core"),        # centripetal force keeps circular motion
        118: (T_NLM, "core"),         # friction: KE → heat
        119: (T_NLM, "core"),         # bus turns right → passengers lean left (inertia)
        120: (T_SOUND, "core"),       # sitarist adjusts frequency (tuning)
    },
    "69th BPSC": {
        2: (T_SPACE, "extended"),     # space missions match
        9: (T_SPACE, "extended"),     # pulsars – rotating neutron stars
        14: (T_MODERN, "extended"),   # Manhattan Project
        15: (T_LIGHT1, "core"),       # concave mirror, same-size image → object at C
        16: (T_MODERN, "core"),       # photoelectric cell: light → electrical energy
        19: (T_ELEC, "core"),         # all three liquids conduct (none is a bad conductor)
        21: (T_GRAV, "core"),         # free fall on Moon – same velocity at any instant
        26: (T_MAG, "core"),          # AC produced by dynamo
        27: (T_ELEC, "core"),         # current density – vector
        40: (T_LIGHT1, "core"),       # colour mixing: magenta, teal, mauve, cyan
    },
    "70th BPSC": {
        6: (T_UNITS, "core"),         # odometer – distance
        15: (T_HEAT, "core"),         # radiation needs no medium
        20: (T_HEAT, "extended"),     # liquefaction: low T, high P
        45: (T_WAVES, "core"),        # infrared for body aches
        70: (T_GRAV, "extended"),     # comet – highly elongated elliptical orbit
        85: (T_NLM, "core"),          # non-contact force – magnetic
        104: (T_ELEC, "core"),        # R = V/I
        113: (T_THERMO, "core"),      # solar constant 1.4 kW/m²
        123: (T_MAG, "core"),         # tesla = N/(A·m)
        150: (T_MODERN, "extended"),  # first atomic model – J. J. Thomson
    },
    "71st BPSC (2025)": {
        121: (T_NLM, "core"),         # friction = 200 N at constant velocity
        122: (T_NLM, "core"),         # F = ma; a = 2s/t² = 2 m/s²; 7 t → 14000 N
        123: (T_POWER, "core"),       # 250 units = 250 kWh = 9 × 10⁸ J
        124: (T_SOUND, "core"),       # sound fastest in steel
        125: (T_NLM, "core"),         # falling body – energy conservation not violated
        126: (T_NLM, "core"),         # W = 140 × 15 = 2100 J
        127: (T_LIGHT1, "core"),      # focal length of plane mirror – infinity
        128: (T_POWER, "core"),       # IR² is not power
        129: (T_POWER, "core"),       # short circuit – current increases very high
        130: (T_POWER, "core"),       # fuse protects equipment
    },
}

# (exam, q_no) -> (study answer, status, note). Used where the workbook has no
# key (71st) or where the key deserves a visible editorial remark.
STUDY_OVERRIDES: dict[tuple[str, int], tuple[str, str, str]] = {
    ("71st BPSC (2025)", 121): ("200 Newtons", "added-answer",
        "No key in workbook. Constant velocity ⇒ net force zero ⇒ friction balances the 200 N push."),
    ("71st BPSC (2025)", 122): ("14000 Newtons", "added-answer",
        "No key in workbook. s = ½at² ⇒ a = 2×400/20² = 2 m/s²; F = ma = 7000 kg × 2 = 14 000 N."),
    ("71st BPSC (2025)", 123): ("9x10⁸", "added-answer",
        "No key in workbook. 1 unit = 1 kWh = 3.6 × 10⁶ J ⇒ 250 units = 9 × 10⁸ J."),
    ("71st BPSC (2025)", 124): ("Steel", "added-answer",
        "No key in workbook. v(solid) > v(liquid) > v(gas); steel ≈ 5,000–6,000 m/s."),
    ("71st BPSC (2025)", 125): ("The principle of conservation of energy is not violated", "added-answer",
        "No key in workbook. Lost PE appears as KE; total mechanical energy is conserved."),
    ("71st BPSC (2025)", 126): ("2100 Joule", "added-answer",
        "No key in workbook. W = F·s = 140 N × 15 m = 2100 J."),
    ("71st BPSC (2025)", 127): ("Infinity", "added-answer",
        "No key in workbook (option A is garbled in the source). Plane mirror: R = ∞ ⇒ f = R/2 = ∞; power zero."),
    ("71st BPSC (2025)", 128): ("IR²", "added-answer",
        "No key in workbook. P = VI = I²R = V²/R; IR² is not a power expression."),
    ("71st BPSC (2025)", 129): ("Increases very high", "added-answer",
        "No key in workbook. Short circuit ⇒ resistance ≈ 0 ⇒ I = V/R becomes very large (fuse melts)."),
    ("71st BPSC (2025)", 130): ("Fuse", "added-answer",
        "No key in workbook. Among the options the fuse is the protective device (strictly, earthing protects people from shock; the fuse protects equipment from over-current)."),
    ("60-62nd BPSC (Combined)", 50): ("Phosphorus-32", "editor-corrected",
        "Workbook key says (d) Cobalt-60. P-32 (β-emitter) is the isotope used in leukaemia/polycythaemia therapy; Co-60 is an external γ source. Study answer follows the standard reference."),
    ("65th BPSC", 112): ("None of the above/More than one of the above", "source-key",
        "Workbook key is (E). Light actually takes ≈ 8 min 20 s, which is why the bare '8 minutes' option was not accepted by the key; remember the value as 8 min 20 s (≈ 500 s)."),
    ("68th BPSC", 112): ("None of the above", "source-key",
        "Key is (E): shrillness (pitch) is decided by frequency, which is not among options A–C."),
    ("69th BPSC", 19): ("None of the above", "source-key",
        "Key is (D): salted water, orange juice and lemon juice are all electrolytes, so none is a bad conductor."),
}


def load_rows() -> dict[tuple[str, int], dict]:
    rows: dict[tuple[str, int], dict] = {}
    wb = load_workbook(PHY / "56-59th BPSC Preliminary Examination Question Paper.xlsx",
                       data_only=True, read_only=True)
    ws = wb.active
    ws.calculate_dimension(force=True)
    for r in ws.iter_rows(min_row=5, max_row=180, values_only=True):
        if isinstance(r[0], (int, float)) and r[1]:
            rows[("56-59th BPSC (2015)", int(r[0]))] = {
                "question": str(r[1]),
                "options": [str(x) if x is not None else "" for x in r[2:6]],
                "source_key": str(r[6] or ""),
            }
    wb = load_workbook(BIO / "BPSC_Master_Question_Bank_60_to_71.xlsx", data_only=True, read_only=True)
    ws = wb["Master_All_Questions"]
    for r in ws.iter_rows(min_row=4, values_only=True):
        if r[0] and isinstance(r[1], (int, float)) and r[2]:
            rows[(str(r[0]), int(r[1]))] = {
                "question": str(r[2]),
                "options": [str(x) if x is not None else "" for x in r[3:8]],
                "source_key": str(r[9] or r[8] or ""),
            }
    return rows


def resolve_answer(raw: str, options: list[str]) -> str:
    s = raw.strip()
    m = re.fullmatch(r"\(?([A-Ea-e])\)?", s)
    if m:
        i = ord(m.group(1).upper()) - ord("A")
        return options[i] if 0 <= i < len(options) and options[i] else s.upper()
    return s


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    source = load_rows()
    records, missing = [], []
    for exam in EXAM_ORDER:
        for qno, (topic, scope) in sorted(SELECTION[exam].items()):
            row = source.get((exam, qno))
            if not row:
                missing.append((exam, qno)); continue
            source_answer = resolve_answer(row["source_key"], row["options"]) if row["source_key"] else ""
            if (exam, qno) in STUDY_OVERRIDES:
                study_answer, status, note = STUDY_OVERRIDES[(exam, qno)]
            else:
                study_answer, status, note = source_answer, "source-key", "Answer follows the verified answer field of the supplied workbook."
            records.append({
                "exam": exam, "q_no": qno, "scope": scope, "topic": topic,
                "recent": "yes" if exam in RECENT else "no",
                "question": " ".join(row["question"].split()),
                "options": " | ".join(x for x in row["options"] if x),
                "source_key_raw": row["source_key"], "source_answer": source_answer,
                "study_answer": study_answer, "status": status, "editor_note": note,
            })
    if missing:
        raise SystemExit(f"Missing selected rows: {missing}")

    fields = list(records[0])
    with (OUT / "pyq_analysis.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, lineterminator="\n")
        w.writeheader(); w.writerows(records)

    topic_all = Counter(r["topic"] for r in records)
    topic_core = Counter(r["topic"] for r in records if r["scope"] == "core")
    topic_recent = Counter(r["topic"] for r in records if r["recent"] == "yes")
    papers = defaultdict(set)
    recent_papers = defaultdict(set)
    for r in records:
        papers[r["topic"]].add(r["exam"])
        if r["recent"] == "yes":
            recent_papers[r["topic"]].add(r["exam"])
    # score: recent questions count double, plus breadth across papers
    def score(t):
        return topic_all[t] + topic_recent[t] + 0.5 * len(papers[t])
    ranking = []
    for rank, t in enumerate(sorted(topic_all, key=lambda t: (-score(t), t)), 1):
        n, nr, np_, nrp = topic_all[t], topic_recent[t], len(papers[t]), len(recent_papers[t])
        if nr >= 6 or (n >= 9 and nrp >= 3):
            pr = "A"
        elif nr >= 3 or n >= 6:
            pr = "B"
        else:
            pr = "C"
        ranking.append({"rank": rank, "topic": t, "all_questions": n, "core_questions": topic_core[t],
                        "recent_questions": nr, "papers_present": np_, "recent_papers_present": nrp,
                        "score": round(score(t), 1), "priority": pr})
    summary = {
        "method": ("Manual question-by-question classification of all 1,623 questions in the two "
                   "supplied XLSX files (56-59th … 71st BPSC). 'recent' = 66th–71st (last six papers); "
                   "score = all + recent + 0.5 × papers-present."),
        "total_questions": len(records),
        "recent_questions": sum(1 for r in records if r["recent"] == "yes"),
        "scope_counts": dict(Counter(r["scope"] for r in records)),
        "exam_counts": {e: sum(1 for r in records if r["exam"] == e) for e in EXAM_ORDER},
        "recent_exams": RECENT,
        "topic_ranking": ranking,
        "status_counts": dict(Counter(r["status"] for r in records)),
    }
    (OUT / "pyq_summary.json").write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {len(records)} reviewed physics questions "
          f"({summary['recent_questions']} from the last six papers)")
    for r in ranking:
        print(f"{r['rank']:2}. {r['topic']:<62} all {r['all_questions']:2} | recent {r['recent_questions']:2} "
              f"| papers {r['papers_present']:2} | {r['priority']}")
    for e in EXAM_ORDER:
        print(f"   {e:<26} {summary['exam_counts'][e]:2}")


if __name__ == "__main__":
    main()
